#! -*-conding:utf-8 -*-
#@Time: 2020/6/6 0006 16:39
#@swzhou
'''
读取、操作excel
'''

import openpyxl
import os
import sys
base_path = os.path.dirname(os.path.dirname(__file__))
from Router.Common.handle_ini import handleIni

file = handleIni.get_value('file_case','Config_file')
file_path = base_path + file

class HandleExcel:

    def load_excel(self):
        '''
        加载指定excel文件
        :param file: 文件路径
        :return: open对象
        '''
        
        open = openpyxl.open(file_path)
        return open

    def get_sheet_data(self,index = None):
        '''
        获取excel中指定sheet
        :param index: sheet number
        :return: sheet
        '''
        if index == None:
            index = 0
        sheet = self.load_excel().sheetnames
        sheet_data = self.load_excel()[sheet[index]]
        return sheet_data

    def get_cell_value(self,row,col,index = None):
        '''
        获取指定单元格数据
        :param row: 行
        :param col: 列
        :param index: 第几个sheet
        :return: 单元格数据
        '''
        if index == None:
            index = 0
        value = self.get_sheet_data(index).cell(row=row,column=col).value
        return value

    def get_rows(self,index = None):
        '''
        获取最大行数
        :param index: sheet num
        :return:
        '''
        if index == None:
            index = 0
        row = self.get_sheet_data(index).max_row
        return row

    def get_row_value(self,row,index = None):
        '''
        获取指定行内容,默认取第一个sheet
        :param row: 
        :param index: 
        :return: 
        '''
        if index == None:
            index = 0
        row_list = []
        row_value = self.get_sheet_data(index)[row]
        for v in row_value:
            row_list.append(v.value)
        return row_list

    def write_data(self,row,col,value):
        '''
        单元格写入指定数据
        :param row: 行
        :param col: 列
        :param value: 数据值
        '''
        wb = self.load_excel()
        wr = wb.active
        wr.cell(row,col,value)
        wb.save(file_path)
        
    def get_columns_value(self,key = None):
        '''
        获取某列的值
        :param key:某一列 “A” 
        :return: 列值的list 
        '''
        columns_list = []
        if key == None:
            key = "A"
        columns_list_data = self.get_sheet_data()[key]
        for i in columns_list_data:
            columns_list.append(i.value)
        return columns_list
        
    def get_rows_number(self,case_id):
        '''
        通过case_id 判断当前行号
        :param case_id: 
        :return:  int
        '''
        cols_data = self.get_columns_value()
        num = 1
        for col_data in cols_data:
            if case_id == col_data:
                return num
            num += 1
        return num
    
    def get_excel_data(self):
        '''
        获取excel所有行数据 -默认第一个sheet
        :return: 
        '''
        data_list = []
        rows = self.get_rows()
        for i in range(rows):
            data_list.append(self.get_row_value(i + 2))
        data_list.pop()# +2会多出一行空数据
        return data_list

handleExcel = HandleExcel()

if __name__ == '__main__':
    # print(handleExcel.get_cell_value(1,3))
    # print(handleExcel.get_row_value(1))
    # print(handleExcel.get_rows())
    # handleExcel.write_data(2,13,'通过')
    print(handleExcel.get_excel_data())
        
